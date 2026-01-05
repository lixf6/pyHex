"""
Excel to CFG Converter Tool

将Excel文件（xls/xlsx）转换为CANape配置文件（cfg）格式。

Excel格式要求：
- 第一行固定为：名称、角色、10msRStr、100msRStr、Polling_100ms、Polling_500ms、Polling_1s
- 从第二行开始，每行包含信号名称和对应的X标记
- X在"10msRStr"列表示10ms周期（cfg中为0）
- X在"100msRStr"列表示100ms周期（cfg中为1）

CFG格式：
- 前3行为固定格式（注释和版本信息）
- 第4行为空行
- 从第5行开始，每行格式为：信号名称;周期值
"""

import os
import logging
import tempfile
import time
import xml.etree.ElementTree as ET
from typing import List, Tuple, Optional

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExcelToCfgConverter:
    """Excel转CFG工具类"""
    
    # CFG文件固定头部
    CFG_HEADER = [
        "// Selection file",
        "// generated with : Vector CANape x64 Version 17.0.70.921",
        "* cfg-file-version : 1.0",
        "",  # 空行
    ]
    
    # Excel列名映射（期望的列名）
    EXPECTED_COLUMNS = ["名称", "角色", "10msRStr", "100msRStr", "Polling_100ms", "Polling_500ms", "Polling_1s"]
    
    def __init__(self, excel_path: str):
        """
        初始化转换器
        
        Args:
            excel_path: Excel文件路径（支持.xls和.xlsx）
        """
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel文件不存在: {excel_path}")
        
        self.excel_path = excel_path
        self.file_ext = os.path.splitext(excel_path)[1].lower()
        
        if self.file_ext not in ['.xls', '.xlsx']:
            raise ValueError(f"不支持的文件格式: {self.file_ext}，仅支持.xls和.xlsx")
        
        # 检查库的可用性（在运行时检查，因为模块导入时可能还未安装）
        global OPENPYXL_AVAILABLE, XLRD_AVAILABLE
        try:
            import openpyxl  # type: ignore[import]
            OPENPYXL_AVAILABLE = True
        except ImportError:
            OPENPYXL_AVAILABLE = False
        
        # 动态检查 xlrd 是否可用（不再强制校验版本号，兼容已安装的环境）
        try:
            import xlrd as _xlrd  # noqa: F401
            XLRD_AVAILABLE = True
        except ImportError:
            XLRD_AVAILABLE = False
        
        # 优先使用openpyxl，因为它可以读取.xlsx和XML格式的.xls文件
        if not OPENPYXL_AVAILABLE:
            if self.file_ext == '.xlsx':
                raise ImportError("需要安装openpyxl库来处理.xlsx文件: pip install openpyxl")
            elif self.file_ext == '.xls':
                # 对于.xls文件，先尝试openpyxl（支持XML格式），如果失败再提示安装xlrd
                if not XLRD_AVAILABLE:
                    raise ImportError("需要安装openpyxl或xlrd库来处理.xls文件: pip install openpyxl 或 pip install xlrd")
    
    def _is_xml_format(self) -> bool:
        """检测文件是否为XML格式"""
        try:
            with open(self.excel_path, 'rb') as f:
                header_bytes = f.read(200)
                header = header_bytes.decode('utf-8', errors='ignore')
                # 去掉 UTF-8 BOM 和首尾空白
                header = header.lstrip('\ufeff').strip()
                return header.startswith('<?xml') or header.startswith('<Workbook')
        except Exception:
            return False
    
    def _read_excel_xml(self) -> List[List[str]]:
        """直接解析XML格式的Excel文件"""
        try:
            tree = ET.parse(self.excel_path)
            root = tree.getroot()
            
            # Excel XML命名空间
            ns = {
                'ss': 'urn:schemas-microsoft-com:office:spreadsheet',
                'html': 'http://www.w3.org/TR/REC-html40'
            }
            
            # 查找第一个Worksheet
            worksheet = root.find('.//ss:Worksheet', ns)
            if worksheet is None:
                raise ValueError("未找到Worksheet元素")
            
            # 查找Table
            table = worksheet.find('ss:Table', ns)
            if table is None:
                raise ValueError("未找到Table元素")
            
            data = []
            for row in table.findall('ss:Row', ns):
                # 处理带有Index属性的单元格
                row_data = []
                current_col = 0
                
                for cell in row.findall('ss:Cell', ns):
                    # 检查是否有Index属性（从1开始）
                    index_attr = cell.get('{urn:schemas-microsoft-com:office:spreadsheet}Index')
                    if index_attr:
                        target_col = int(index_attr) - 1  # 转换为0基索引
                        # 填充中间的空列
                        while len(row_data) < target_col:
                            row_data.append("")
                        current_col = target_col
                    
                    # 获取单元格值
                    data_elem = cell.find('ss:Data', ns)
                    if data_elem is not None:
                        cell_value = data_elem.text if data_elem.text else ""
                    else:
                        cell_value = ""
                    
                    # 如果当前列已经有数据，扩展列表
                    if current_col < len(row_data):
                        row_data[current_col] = cell_value
                    else:
                        row_data.append(cell_value)
                    
                    current_col += 1
                
                data.append(row_data)
            
            return data
        except ET.ParseError as e:
            raise ValueError(f"XML解析失败: {e}")
    
    def _read_excel_xlsx(self) -> List[List[str]]:
        """读取.xlsx格式的Excel文件"""
        # 如果是.xls文件但实际是XML格式，先尝试直接解析XML
        if self.file_ext == '.xls' and self._is_xml_format():
            try:
                return self._read_excel_xml()
            except Exception as e:
                logger.warning(f"XML解析失败，尝试使用openpyxl: {e}")
                # 如果XML解析失败，尝试将文件复制为.xlsx格式让openpyxl读取
                try:
                    with tempfile.NamedTemporaryFile(mode='w+b', suffix='.xlsx', delete=False) as tmp:
                        with open(self.excel_path, 'rb') as src:
                            tmp.write(src.read())
                        tmp_path = tmp.name
                    
                    try:
                        wb = openpyxl.load_workbook(tmp_path, data_only=True)
                        ws = wb.active
                        data = []
                        for row in ws.iter_rows(values_only=True):
                            row_data = [str(cell) if cell is not None else "" for cell in row]
                            data.append(row_data)
                        return data
                    finally:
                        if os.path.exists(tmp_path):
                            os.unlink(tmp_path)
                except Exception:
                    raise ValueError(f"无法读取Excel文件: {e}")
        
        # 正常读取.xlsx文件
        try:
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        except Exception as e:
            if self.file_ext == '.xls':
                # 对于旧版二进制 .xls 文件，openpyxl 无法处理，提示用户使用 xlrd 或转为 .xlsx
                raise ValueError(
                    "无法读取Excel文件：检测到旧版 .xls 格式，openpyxl 不支持此格式；"
                    "请安装 xlrd==1.2.0 并重试，或在 Excel 中将文件另存为 .xlsx 后再导入。\n"
                    f"底层错误信息: {e}"
                )
            raise
        
        # 获取第一个工作表
        ws = wb.active
        
        data = []
        for row in ws.iter_rows(values_only=True):
            # 将None转换为空字符串，并转换为字符串列表
            row_data = [str(cell) if cell is not None else "" for cell in row]
            data.append(row_data)
        
        return data
    
    def _read_excel_xls(self) -> List[List[str]]:
        """读取.xls格式的Excel文件"""
        wb = xlrd.open_workbook(self.excel_path)
        # 获取第一个工作表
        ws = wb.sheet_by_index(0)
        
        data = []
        for row_idx in range(ws.nrows):
            row_data = []
            for col_idx in range(ws.ncols):
                cell_value = ws.cell_value(row_idx, col_idx)
                row_data.append(str(cell_value) if cell_value else "")
            data.append(row_data)
        
        return data
    
    def _read_excel(self) -> List[List[str]]:
        """读取Excel文件内容"""
        # 优先尝试使用openpyxl（支持.xlsx和XML格式的.xls）
        if OPENPYXL_AVAILABLE:
            try:
                return self._read_excel_xlsx()
            except Exception as e:
                # 如果openpyxl读取失败且是.xls文件，尝试使用xlrd
                if self.file_ext == '.xls':
                    if XLRD_AVAILABLE:
                        logger.warning(f"openpyxl读取 .xls 文件失败，尝试使用 xlrd: {e}")
                        return self._read_excel_xls()
                    # 没有可用的 xlrd，给出清晰的提示
                    raise ValueError(
                        "无法读取旧版 .xls 格式的 Excel 文件：openpyxl 不支持该格式，"
                        "且未检测到可用的 xlrd 旧版本 (需 xlrd==1.2.0)。"
                        "请安装 xlrd==1.2.0 或在 Excel 中将文件另存为 .xlsx 后再导入。\n"
                        f"底层错误信息: {e}"
                    )
                raise
        else:
            # 如果没有openpyxl，只能使用xlrd读取.xls
            if self.file_ext == '.xls':
                return self._read_excel_xls()
            else:
                raise ImportError("需要安装openpyxl库来处理.xlsx文件: pip install openpyxl")
    
    def _normalize_cell_value(self, value: str) -> str:
        """标准化单元格值，用于匹配X标记"""
        if not value:
            return ""
        # 去除空格并转换为大写，用于匹配X
        # 处理可能的None值
        if value is None:
            return ""
        return str(value).strip().upper()
    
    def _find_column_indices(
        self,
        header_row: List[str],
    ) -> Tuple[
        Optional[int],  # 10msRStr
        Optional[int],  # 100msRStr
        Optional[int],  # Polling_100ms
        Optional[int],  # Polling_500ms
        Optional[int],  # Polling_1s
    ]:
        """
        查找周期相关列的索引
        
        Returns:
            (col_10ms_idx, col_100ms_idx, col_poll_100_idx, col_poll_500_idx, col_poll_1s_idx)
        """
        col_10ms_idx = None
        col_100ms_idx = None
        col_poll_100_idx = None
        col_poll_500_idx = None
        col_poll_1s_idx = None
        
        # 标准化表头行
        normalized_header = [self._normalize_cell_value(cell) for cell in header_row]
        
        # 查找列索引
        for idx, col_name in enumerate(normalized_header):
            if "10MSRSTR" in col_name or col_name == "10MSRSTR":
                col_10ms_idx = idx
            elif "100MSRSTR" in col_name or col_name == "100MSRSTR":
                col_100ms_idx = idx
            elif "POLLING_100MS" in col_name or col_name == "POLLING_100MS":
                col_poll_100_idx = idx
            elif "POLLING_500MS" in col_name or col_name == "POLLING_500MS":
                col_poll_500_idx = idx
            elif "POLLING_1S" in col_name or col_name == "POLLING_1S":
                col_poll_1s_idx = idx
        
        return (
            col_10ms_idx,
            col_100ms_idx,
            col_poll_100_idx,
            col_poll_500_idx,
            col_poll_1s_idx,
        )
    
    def _find_signal_name_column(self, header_row: List[str]) -> Optional[int]:
        """查找信号名称列的索引（通常是第一列"名称"）"""
        normalized_header = [self._normalize_cell_value(cell) for cell in header_row]
        
        # 查找"名称"列
        for idx, col_name in enumerate(normalized_header):
            if "名称" in col_name or col_name == "名称" or col_name == "NAME":
                return idx
        
        # 如果找不到，默认使用第一列
        return 0
    
    def parse_excel(self) -> Tuple[List[Tuple[str, int]], int]:
        """
        解析Excel文件，提取信号名称和周期
        
        Returns:
            Tuple[List[Tuple[str, int]], int]: ([(信号名称, 周期值), ...], Excel中的总信号数)
            周期值：0表示10ms，1表示100ms
            总信号数：Excel中所有有效的信号名称数量（包括没有周期标记的）
        """
        data = self._read_excel()
        
        if len(data) < 2:
            raise ValueError("Excel文件至少需要包含表头和数据行")
        
        header_row = data[0]
        signal_name_col = self._find_signal_name_column(header_row)
        (
            col_10ms_idx,
            col_100ms_idx,
            col_poll_100_idx,
            col_poll_500_idx,
            col_poll_1s_idx,
        ) = self._find_column_indices(header_row)
        
        # 至少要有 10msRStr 或 100msRStr 其中之一
        if col_10ms_idx is None and col_100ms_idx is None:
            raise ValueError("未找到10msRStr或100msRStr列，请检查Excel文件格式")
        
        signals = []
        total_signals_count = 0  # Excel中的总信号数
        
        # 从第二行开始处理数据
        for row_idx, row in enumerate(data[1:], start=2):
            if not row:
                continue
            
            # 获取信号名称
            if signal_name_col is None or signal_name_col >= len(row):
                continue
            
            signal_name_raw = row[signal_name_col] if signal_name_col < len(row) else ""
            # 保持原始信号名称的大小写，只去除首尾空格
            signal_name = str(signal_name_raw).strip() if signal_name_raw else ""
            
            # 跳过空行或无效信号名
            if not signal_name or signal_name.upper() == "NONE" or signal_name.upper() == "NAN":
                continue
            
            # 统计Excel中的总信号数（只要信号名有效就计数）
            total_signals_count += 1
            
            # 周期判定规则：
            # - 如果 10msRStr 列有 X，则为 0（10ms）
            # - 否则，只要 100msRStr / Polling_100ms / Polling_500ms / Polling_1s 任意一列有 X，则为 1（100ms）
            period = None
            
            # 10msRStr
            if col_10ms_idx is not None and col_10ms_idx < len(row):
                cell_value = self._normalize_cell_value(row[col_10ms_idx])
                if cell_value == "X":
                    period = 0  # 10ms
            
            # 100msRStr & Polling_* 统统当作 100ms
            if period is None:
                # 100msRStr
                if col_100ms_idx is not None and col_100ms_idx < len(row):
                    cell_value = self._normalize_cell_value(row[col_100ms_idx])
                    if cell_value == "X":
                        period = 1
                # Polling_100ms
                if period is None and col_poll_100_idx is not None and col_poll_100_idx < len(row):
                    cell_value = self._normalize_cell_value(row[col_poll_100_idx])
                    if cell_value == "X":
                        period = 1
                # Polling_500ms
                if period is None and col_poll_500_idx is not None and col_poll_500_idx < len(row):
                    cell_value = self._normalize_cell_value(row[col_poll_500_idx])
                    if cell_value == "X":
                        period = 1
                # Polling_1s
                if period is None and col_poll_1s_idx is not None and col_poll_1s_idx < len(row):
                    cell_value = self._normalize_cell_value(row[col_poll_1s_idx])
                    if cell_value == "X":
                        period = 1
            
            # 如果找到了周期标记，添加到结果中
            if period is not None:
                signals.append((signal_name, period))
            # else:
            #     # 如果两个周期列都没有X，记录警告但继续处理
            #     logger.warning(f"第{row_idx}行的信号'{signal_name}'未找到周期标记（10msRStr或100msRStr列中的X）")
        
        return signals, total_signals_count
    
    def convert_to_cfg(self, output_path: Optional[str] = None) -> str:
        """
        将Excel转换为CFG格式文件
        
        Args:
            output_path: 输出文件路径，如果为None则自动生成
        
        Returns:
            输出文件路径
        """
        # 记录转换开始时间
        start_time = time.time()
        start_time_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(start_time))
        
        # 解析Excel文件
        signals, total_excel_signals = self.parse_excel()
        signals_with_period = len(signals)
        signals_without_period = max(total_excel_signals - signals_with_period, 0)
        
        # 转换前：打印Excel中的信号数统计
        print(f"[转换前] Excel文件中找到 {total_excel_signals} 个有效信号")
        print(f"[转换前] 其中 {signals_with_period} 个信号存在周期标记（有 X，将被转换）")
        print(f"[转换前] 其中 {signals_without_period} 个信号不存在周期标记（无 X，不转换）")
        print(f"[转换开始时间] {start_time_str}")
        
        if not signals:
            raise ValueError("未找到任何有效的信号数据（没有找到带周期标记的信号）")
        
        # 生成输出文件路径
        if output_path is None:
            base_name = os.path.splitext(os.path.basename(self.excel_path))[0]
            output_dir = os.path.dirname(self.excel_path) or "."
            output_path = os.path.join(output_dir, f"{base_name}.cfg")
        
        # 写入CFG文件
        written_signals_count = 0
        with open(output_path, 'w', encoding='utf-8') as f:
            # 写入固定头部
            for line in self.CFG_HEADER:
                f.write(line + '\n')
            
            # 写入信号数据
            for signal_name, period in signals:
                f.write(f"{signal_name};{period}\n")
                written_signals_count += 1
        
        # 记录转换结束时间
        end_time = time.time()
        end_time_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(end_time))
        elapsed_time = end_time - start_time
        
        # 转换后：验证并打印写入的信号数
        print(f"[转换后] CFG文件中成功写入 {written_signals_count} 个信号")
        print(f"[转换结束时间] {end_time_str}")
        print(f"[转换耗时] {elapsed_time:.3f} 秒")
        
        # 验证转换过程中没有丢失信号
        if written_signals_count != len(signals):
            error_msg = (
                f"信号数量不匹配！解析得到 {len(signals)} 个信号，但只写入了 {written_signals_count} 个信号。"
                f"转换过程中可能丢失了 {len(signals) - written_signals_count} 个信号！"
            )
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        # 验证成功
        print(f"[验证成功] 转换过程中没有丢失信号，所有 {written_signals_count} 个信号都已成功写入CFG文件")
        
        logger.info(f"成功转换Excel文件，生成CFG文件: {output_path}，共{written_signals_count}个信号，耗时{elapsed_time:.3f}秒")
        
        return output_path


def convert_excel_to_cfg(excel_path: str, output_path: Optional[str] = None) -> str:
    """
    便捷函数：将Excel文件转换为CFG文件
    
    Args:
        excel_path: Excel文件路径（支持.xls和.xlsx）
        output_path: 输出CFG文件路径，如果为None则自动生成
    
    Returns:
        输出文件路径
    
    Example:
        >>> convert_excel_to_cfg("xcp信号测试.xls", "output.cfg")
        'output.cfg'
        
    Usage:
        # 基本使用
        from utils.excel_to_cfg_converter import convert_excel_to_cfg
        
        # 转换Excel文件，自动生成输出文件名
        output_file = convert_excel_to_cfg("xcp信号测试.xls")
        
        # 指定输出文件路径
        output_file = convert_excel_to_cfg("xcp信号测试.xls", "output.cfg")
        
        # 使用类进行更精细的控制
        from utils.excel_to_cfg_converter import ExcelToCfgConverter
        
        converter = ExcelToCfgConverter("xcp信号测试.xls")
        signals, total_count = converter.parse_excel()  # 获取解析的信号列表和总数
        output_file = converter.convert_to_cfg("output.cfg")
    """
    converter = ExcelToCfgConverter(excel_path)
    return converter.convert_to_cfg(output_path)


if __name__ == "__main__":
    """
    测试脚本：使用当前目录下的 xcp信号测试.xls 文件进行测试
    """
    import sys
    
    # 测试文件路径
    # test_excel_file = "xcp信号测试.xls"
    # test_excel_file = "test1223_1.xls"
    # test_excel_file = "test1223_6.xls"
    # test_excel_file = "test_1223_7.xls"
    # test_excel_file = "test1223.xls"
    test_excel_file = "test_Original.xls"
    # test_excel_file = "1223_original_1.xls"
    
    # 检查文件是否存在
    if not os.path.exists(test_excel_file):
        print(f"错误: 找不到测试文件: {test_excel_file}")
        print(f"请确保文件存在于当前目录: {os.getcwd()}")
        sys.exit(1)
    
    # 自动生成输出文件名：使用输入文件名，后缀改为.cfg
    # 如果未指定输出文件，函数会自动生成，这里预先计算用于显示
    base_name = os.path.splitext(test_excel_file)[0]
    expected_output_file = f"{base_name}.cfg"
    
    print("=" * 60)
    print("Excel转CFG转换工具测试")
    print("=" * 60)
    print(f"输入文件: {test_excel_file}")
    print(f"输出文件: {expected_output_file} (自动生成)")
    print("-" * 60)
    
    try:
        # 执行转换（不指定输出文件，让函数自动生成）
        output_file = convert_excel_to_cfg(test_excel_file)
        
        print("-" * 60)
        print(f"[OK] 测试成功！输出文件: {output_file}")
        
        # 显示输出文件的前几行
        print("\n输出文件前20行预览:")
        print("-" * 60)
        with open(output_file, 'r', encoding='utf-8') as f:
            for i, line in enumerate(f, 1):
                if i <= 20:
                    print(f"{i:3d}: {line.rstrip()}")
                else:
                    break
        
        print("\n" + "=" * 60)
        print("测试完成！")
        print("=" * 60)
        
    except Exception as e:
        print("-" * 60)
        print(f"[ERROR] 测试失败: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

